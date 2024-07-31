import openpyxl
from datetime import datetime
import os
from openpyxl.drawing.image import Image
from openpyxl import load_workbook


# 1.請求書のダミーデータを読み込む
# load_workbookは、既に存在するファイルを読み込むときに使う
wb = openpyxl.load_workbook("files/invoice_data.xlsx", data_only=True)
ws = wb.active

# valuesにワークシートの全てのセルの値をリストとして取得
values = list(ws.values)

lastrow = len(values)

# 2.請求書のテンプレートを読み込む
wb = openpyxl.load_workbook("files/invoice.xlsx", data_only=True)
ws = wb.active

# 3.日付の取得やフォルダの作成

# 現在の日付時刻を取得
current_date = datetime.now()

# 日付のみに整形（2024年7月31日)
invoice_date = current_date.strftime("%Y年%m月&d日")

# 番号のみに整形（202407）
year_month = current_date.strftime("%Y%m")

# 件名用に整形（月のみになる）
invoice_month = current_date.month

# 保存先のフォルダを指定 シングルクォーテーションでないといけない？↓
output_folder = f"請求書_{current_date.strftime('%Y年%m月')}"

# フォルダ作成 exist_ok==True必須
os.makedirs(output_folder, exist_ok=True)

# 請求書ファイルのpathを作成
output_file = f"{output_folder}/請求書_{current_date.strftime('%Y年%m月')}.xlsx"

# 請求書番号の初期化（毎月リセット）
invoice_number = 1

# 4.繰り返し処理で請求書テンプレートに値を入れていく

for index in range(lastrow):
    # 1行目がヘッダーだから、取得しないようにする
    if not index == 0:
        # 12行目が請求金額であるため、そこに値が存在しないため、スキップする（繰り替え処理をしない）
        if values[index][12] is None:
            continue

        # 会社名を取得ここから↓valueからすべてを持ってきている
        sheet_name = str(values[index][0])
        # ワークシートを新しく作成
        copy_ws = wb.copy_worksheet(ws)
        # ワークシートを会社名に変更
        copy_ws.title = sheet_name
        # 会社名（A2がそうだから)
        copy_ws["A2"].value = sheet_name
        # 担当者名(invoice...のA4がそうで、invoice_data....の10行目がそれだから)
        copy_ws["A4"].value = values[index][10]
        # 件名（B7がそう)
        copy_ws["B7"].value = f"{invoice_month}月分請求書"
        # 番号（N2がそうで、上からあらかじめ取得しておいたyear_monthとinvoice_numberがここで使われる） 03dをつけることによって001 002 003のような形になる
        copy_ws["N2"].value = f"{year_month}-{invoice_number:03d}"
        invoice_number += 1
        # 請求日(N3がそう)
        copy_ws["N3"].value = invoice_date

        # 摘要欄
        copy_ws["A14"].value = values[index][13]
        copy_ws["A15"].value = values[index][18]
        copy_ws["A16"].value = values[index][23]
        copy_ws["A17"].value = values[index][28]
        copy_ws["A18"].value = values[index][33]
        copy_ws["A19"].value = values[index][38]
        copy_ws["A20"].value = values[index][43]
        copy_ws["A21"].value = values[index][48]
        copy_ws["A22"].value = values[index][53]
        copy_ws["A23"].value = values[index][58]

        # 数量
        copy_ws["J14"].value = values[index][14]
        copy_ws["J15"].value = values[index][19]
        copy_ws["J16"].value = values[index][24]
        copy_ws["J17"].value = values[index][29]
        copy_ws["J18"].value = values[index][34]
        copy_ws["J19"].value = values[index][39]
        copy_ws["J20"].value = values[index][44]
        copy_ws["J21"].value = values[index][49]
        copy_ws["J22"].value = values[index][54]
        copy_ws["J23"].value = values[index][59]

        # 数量
        copy_ws["K14"].value = values[index][15]
        copy_ws["K15"].value = values[index][20]
        copy_ws["K16"].value = values[index][25]
        copy_ws["K17"].value = values[index][30]
        copy_ws["K18"].value = values[index][35]
        copy_ws["K19"].value = values[index][40]
        copy_ws["K20"].value = values[index][45]
        copy_ws["K21"].value = values[index][50]
        copy_ws["K22"].value = values[index][55]
        copy_ws["K23"].value = values[index][60]

        # 単価
        copy_ws["L14"].value = values[index][16]
        copy_ws["L15"].value = values[index][21]
        copy_ws["L16"].value = values[index][26]
        copy_ws["L17"].value = values[index][31]
        copy_ws["L18"].value = values[index][36]
        copy_ws["L19"].value = values[index][41]
        copy_ws["L20"].value = values[index][46]
        copy_ws["L21"].value = values[index][51]
        copy_ws["L22"].value = values[index][56]
        copy_ws["L23"].value = values[index][61]

        # 金額
        copy_ws["O14"].value = values[index][17]
        copy_ws["O15"].value = values[index][22]
        copy_ws["O16"].value = values[index][27]
        copy_ws["O17"].value = values[index][32]
        copy_ws["O18"].value = values[index][37]
        copy_ws["O19"].value = values[index][42]
        copy_ws["O20"].value = values[index][47]
        copy_ws["O21"].value = values[index][52]
        copy_ws["O22"].value = values[index][57]
        copy_ws["O23"].value = values[index][62]

        # 角印データ貼り付け
        img = Image("files/角印.png")
        img.width = 100
        img.height = 100
        copy_ws.add_image(img, "P5")

wb.save(output_file)
